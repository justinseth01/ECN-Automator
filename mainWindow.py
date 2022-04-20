"""
mainWindow.py

This document is responsible for assembling the layout of the main window
of the UI. It creates the layouts of the varvious group boxes as well as
the button layouts within most of the group boxes. 

This document layout for the main window of the program using pyqt5
it references errorWindow.py to display error windows
as well as settingsWindow.py to display the settings window

Contains:
    -   Constructor with box layout and "settings" "okay" and "cancel" buttons
    -   exit() function that calls unreserveECN and quits program
    -   errorCheck() ensures all info is correctly entered before executing
    -   showSettings() shows the settings window
    -   add() & delete() add/remove a change referencing customWidget.py
    -   reserveECN() returns the number of the next available ECN and
        fills in the name column to "reserve" the ECN
    -   unreserv3eECN() removes info from the "name" column to remove the
        reservation on that row
    -   Group Box functions that set the layout of buttons and assign
        their values to variables

Justin Seth
3/15/2022
"""

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from customWidget import customWidget
from errorWindow import error
from settingsWindow import settingsWindow
import openpyxl
import sys


class mainWindow(QDialog):
    # constructor

    def __init__(self):
        super(mainWindow, self).__init__()
        # Create instance of settings window. Shown in showSettings method
        self.settingsWindow = settingsWindow()

        # okay variable will only be true if all error checks pass
        self.okay = False

        # declare global variables
        self.numChanges = 0
        self.Button1 = []
        self.rows = 4

        # Call method that reads settings.txt and returns array of settings
        self.settings = self.settingsWindow.readSettings()
        """
        settings[0]= first name
        settings[1]= last name
        settings[2]= from email
        settings[3]= to email
        settings[4]= ECN Log Path
        settings[5]= ECN Format Path
        settings[6]= ECN Save Path
        """

        # setting window title to the ecn number
        try:
            # Try calling reserveECN() which returns ECN number and reserves ECN row
            self.ECN = str(self.reserveECN())
        except Exception as e:
            # Show error window with error and exit
            error("Could not open ECN", e)
            sys.exit(0)

        self.setWindowTitle("ECN-"+self.ECN)

        # setting geometry of the window
        self.setGeometry(100, 100, 700, 400)

        # calling the method that create the layout of each group box
        self.createFirstGroup()
        self.createSecondGroup()
        self.createThirdGroup()
        self.createFourthGroup()

        # creating a dialog button for ok and cancel
        self.buttonBox = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

        # Create settings button
        self.settingsButton = QPushButton("Settings")

        # adding action when form is accepted
        self.buttonBox.accepted.connect(self.errorCheck)

        # adding action when form is rejected
        self.buttonBox.rejected.connect(self.exit)

        # Connecting settings button to show settings window
        self.settingsButton.clicked.connect(self.showSettings)

        # creating a vertical layout
        mainLayout = QGridLayout()

        # adding form group box to the layout
        mainLayout.addWidget(self.firstGroup, 0, 0)
        mainLayout.addWidget(self.secondGroup, 1, 0)
        mainLayout.addWidget(self.thirdGroup, 2, 0)
        mainLayout.addWidget(self.fourthGroup, 0, 1, 3, 1)
        mainLayout.addWidget(self.settingsButton, 3, 0)
        mainLayout.addWidget(self.buttonBox, 3, 1)

        # setting lay out
        self.setLayout(mainLayout)

    # unreserves the excel row before quitting the program
    def exit(self):
        print("Exiting")
        self.unreserveECN(self.settings[4], int(self.ECN))
        self.reject
        sys.exit(0)

    # checks information is correct before running program
    def errorCheck(self):
        # if errorCheck stays true, then the rest of the program can run
        errorCheck = True

        # Check project and MM are numbers of at least length 1
        if len(self.mmNumber.text()) > 0:
            try:
                int(self.mmNumber.text())
            except ValueError as detailedMessage:
                error("MM Number is not a number", detailedMessage)
                errorCheck = False
        else:
            error("No MM number entered")
            errorCheck = False

        if len(self.projectNumber.text()) > 0:
            try:
                int(self.projectNumber.text())
            except ValueError as e:
                error("Project Number is not a number", e)
                errorCheck = False
        else:
            error("No project number entered")
            errorCheck = False

        # Make sure rev number is len 1 or less
        if len(self.revNumber.text()) > 1:
            error("Length REV Number > 1")
            errorCheck = False

        if errorCheck:
            self.okay = True
            # Close settings window if open
            if self.settingsWindow.isVisible() == True:
                self.settingsWindow.close()
            self.close()

    # Shows settings window
    def showSettings(self):
        # Only show settings window if it's not already open
        if self.settingsWindow.isVisible() == False:
            self.settingsWindow.show()

    # Adds an additional "change" object when "add" button is pressed
    def add(self):
        self.numChanges += 1  # increase our changes count

        # Add another custom widget row
        self.Button1.append(customWidget())
        self.layout.addWidget(
            QLabel("Change "+str(self.numChanges)), 2*self.numChanges-1, 0)
        self.layout.addWidget(
            self.Button1[self.numChanges-1], 2*self.numChanges, 0, 1, 2)

    # Removes the last "change" object when "delete" button is pressed
    def delete(self):
        if self.numChanges > 0:  # Only delete if there's at least 1 change

            # Delete the last widget row
            self.layout.itemAt(2*self.numChanges).widget().deleteLater()
            self.layout.itemAt(2*self.numChanges+1).widget().deleteLater()
            self.Button1.pop()
            self.numChanges -= 1  # Decrement our changes count

    # Reserves ECN row and returns next available ECN number
    def reserveECN(self):
        workbook = openpyxl.load_workbook(self.settings[4])
        sheet = workbook.active

        currentRow = 1
        currentCol = 2
        cell = sheet.cell(row=currentRow, column=currentCol)

        # find next open ECN for next fully open row
        while True:
            while(cell.value != None):
                currentRow += 1
                cell = sheet.cell(row=currentRow, column=currentCol)
            while(cell.value == None and currentCol < 9):
                currentCol += 1
                cell = sheet.cell(row=currentRow, column=currentCol)
            if(currentCol == 9):
                break
            else:
                currentCol = 2
                currentRow += 1

        sheet["B"+str(currentRow)] = self.settings[0][0].upper()+". " + \
            self.settings[1].capitalize()
        workbook.save(self.settings[4])
        print("Reserving ECN-"+str(sheet.cell(row=currentRow, column=1).value))
        return sheet.cell(row=currentRow, column=1).value

    # Function used to remove the data from the ECN that was reserved
    # Is used anytime the program is canceled
    def unreserveECN(self, ecnLogPath, ecnNumber):
        print("Unreserving ECN")
        path = ecnLogPath

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        # find row with ECN
        currentRow = 1
        cell = sheet.cell(row=currentRow, column=1)
        while(cell.value != ecnNumber):
            currentRow += 1
            cell = sheet.cell(row=currentRow, column=1)

        # Replace reserved ECN cell with empty string
        rows = ["B", "C", "D", "E", "F", "G", "H"]
        for row in rows:
            sheet[row+str(currentRow)] = ""
        workbook.save(ecnLogPath)
        print("ECN removed")

    #########################################################
    ##### methods used to create our different group boxes ##
    #########################################################

    def createFirstGroup(self):
        self.mmNumber = QLineEdit("4329488")
        # self.mmNumber.setStyleSheet("background-color: red;")
        self.partNumber = QLineEdit("ast-20a18-we130")
        self.revNumber = QLineEdit("b")
        self.description = QLineEdit("overall description")
        self.notes = QLineEdit()

        self.firstGroup = QGroupBox("Change Information")

        # creating a form layout
        layout = QFormLayout()

        # adding rows
        layout.addRow(QLabel("MM Number"), self.mmNumber)
        layout.addRow(QLabel("Part Number"), self.partNumber)
        layout.addRow(QLabel("Rev Number"), self.revNumber)
        layout.addRow(QLabel("Description"), self.description)
        layout.addRow(QLabel("Notes"), self.notes)

        self.firstGroup.setLayout(layout)

    def createSecondGroup(self):
        self.secondGroup = QGroupBox("Changes")
        # creating a form layout
        layout = QGridLayout()
        self.changeNames = ["Change 1", "Change 2", "Change 3",
                            "Change 4", "Change 5", "Change 6", "Change 7", "Change 8", "Change 9", "Change 10", "Change 11", "Change 12"]
        self.changes = []
        self.rows = 4

        # Use a for loop and if statements to create the 3 x 4 grid
        # of changes.
        for i, v in enumerate(self.changeNames):
            self.changes.append(QCheckBox(v))
            if i < self.rows:
                layout.addWidget(self.changes[i], i, 0)
            elif i < 2*self.rows:
                layout.addWidget(self.changes[i], i-self.rows, 1)
            else:
                layout.addWidget(self.changes[i], i-2*self.rows, 2)

        self.secondGroup.setLayout(layout)

    def createThirdGroup(self):
        self.projectNumber = QLineEdit("4375932")
        self.serialNumber = QLineEdit("p20-1234")
        self.machine = QLineEdit("MicroFlow 12/14/16")

        self.thirdGroup = QGroupBox("Project Info")
        # creating a form layout
        layout = QFormLayout()

        layout.addRow(QLabel("Serial Number"), self.serialNumber)
        layout.addRow(QLabel("Project Number"), self.projectNumber)
        layout.addRow(QLabel("Machine"), self.machine)

        # setting layout
        self.thirdGroup.setLayout(layout)

    def createFourthGroup(self):
        self.addChange = QPushButton("Add")
        self.addChange.clicked.connect(self.add)

        self.deleteChange = QPushButton("Delete")
        self.deleteChange.clicked.connect(self.delete)

        self.box = QGroupBox("Changes")

        # creating a form layout
        self.layout = QGridLayout()

        self.layout.addWidget(self.addChange, 0, 0)
        self.layout.addWidget(self.deleteChange, 0, 1)
        self.layout.setAlignment(Qt.AlignTop)
        self.add()

        # setting layout
        self.box.setLayout(self.layout)
        self.fourthGroup = QScrollArea()
        self.fourthGroup.setWidget(self.box)
        self.fourthGroup.setWidgetResizable(True)
